#!/usr/bin/env python3
"""
GitRecon v4.0 — GitHub Organisation Security & Repo Analyser
Author : Saurabh Jain

What it does:
  1. Resolves token — GITHUB_TOKEN / GH_TOKEN env vars, then gh CLI fallback
  2. If --orgs not passed → auto-discovers all orgs for the authenticated user
     and runs a full security audit automatically (no-args mode)
  3. Per org, fetches:
       • All members (admins + members) with full profile
       • All public repos under each MEMBER's personal namespace
       • All repos (public + private) under the ORGANISATION itself
       • All STALE BRANCHES (no commit > --stale-days, default 180 days)
         across every org repo, checked concurrently
  4. Security audit (auto in no-args mode, or via --security flag):
       • Members without 2FA  (requires admin:org scope)
       • Default branch protection missing on org repos
       • Org repos with no description
       • Org repos with no license file
       • Dependabot vulnerability alerting disabled
       • Public repos with suspicious/sensitive names
       • Archived repos with open issues still unresolved
       • High admin-to-member ratio warning
       • Stale branch count summary
  5. Exports one XLSX with 6 sheets (or CSV / JSON):
       Summary · Members · Member Repos · Org Repos · Stale Branches · Security Audit

Usage:
  python gitrecon.py                           # auto-discover orgs + full security audit
  python gitrecon.py --orgs mycompany          # target one org
  python gitrecon.py --orgs acme devops        # multiple orgs
  python gitrecon.py --orgs acme --security    # named org + security audit
  python gitrecon.py --orgs acme --stale-days 90
  python gitrecon.py --orgs acme --include-forks --since 2024-01-01
  python gitrecon.py --orgs acme --members-only
  python gitrecon.py --orgs acme --output-format json --no-export

Prerequisites:
  pip install -r requirements.txt
  gh auth login   (or set GITHUB_TOKEN / GH_TOKEN env var)

Scopes needed for full output:
  read:org          — private member listing
  admin:org         — 2FA membership check
  security_events   — Dependabot alert status
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Dict, Iterator, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rich.console import Console
from rich.panel import Panel
from rich.progress import (
    BarColumn, MofNCompleteColumn, Progress,
    SpinnerColumn, TaskProgressColumn, TextColumn, TimeElapsedColumn,
)
from rich.rule import Rule
from rich.table import Table
from rich.traceback import install as install_rich_traceback
from rich import box

install_rich_traceback()

# ── Constants ─────────────────────────────────────────────────────────────────
__version__   = "4.0"
GITHUB_API    = "https://api.github.com"
DATE_FILE     = datetime.now().strftime("%Y-%m-%dT%H-%M")
AUTHOR        = "Saurabh Jain"
STALE_DEFAULT = 180
LOG           = logging.getLogger("gitrecon")
console       = Console()

# Excel colour fills
_HDR       = PatternFill("solid", start_color="1F3864")
_HFNT      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_THIN      = Side(style="thin", color="D9D9D9")
_BDR       = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)
_DFNT      = Font(name="Calibri", size=10)
FILL_RED   = PatternFill("solid", start_color="FFC7CE")
FILL_AMBER = PatternFill("solid", start_color="FFEB9C")
FILL_GREEN = PatternFill("solid", start_color="C6EFCE")
FILL_BLUE  = PatternFill("solid", start_color="DDEBF7")
FILL_GREY  = PatternFill("solid", start_color="D9D9D9")
FILL_WHITE = PatternFill("solid", start_color="FFFFFF")


# ══════════════════════════════════════════════════════════════════════════════
# Data model
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class GHUser:
    login:        str
    name:         str = ""
    email:        str = ""
    company:      str = ""
    location:     str = ""
    public_repos: int = 0
    followers:    int = 0
    created_at:   str = ""
    profile_url:  str = ""
    role:         str = ""    # "admin" | "member"

    @classmethod
    def from_api(cls, data: dict, role: str = "member") -> "GHUser":
        return cls(
            login        = data.get("login", ""),
            name         = data.get("name") or "",
            email        = data.get("email") or "",
            company      = data.get("company") or "",
            location     = data.get("location") or "",
            public_repos = data.get("public_repos", 0),
            followers    = data.get("followers", 0),
            created_at   = (data.get("created_at") or "")[:10],
            profile_url  = data.get("html_url", ""),
            role         = role,
        )


@dataclass
class GHRepo:
    owner:          str
    name:           str
    full_name:      str
    description:    str       = ""
    fork:           bool      = False
    private:        bool      = False
    language:       str       = ""
    stars:          int       = 0
    forks:          int       = 0
    size_kb:        int       = 0
    default_branch: str       = "main"
    topics:         List[str] = field(default_factory=list)
    last_push:      str       = ""
    created_at:     str       = ""
    archived:       bool      = False
    html_url:       str       = ""
    open_issues:    int       = 0

    @classmethod
    def from_api(cls, data: dict) -> "GHRepo":
        return cls(
            owner          = data.get("owner", {}).get("login", ""),
            name           = data.get("name", ""),
            full_name      = data.get("full_name", ""),
            description    = (data.get("description") or "")[:200],
            fork           = data.get("fork", False),
            private        = data.get("private", False),
            language       = data.get("language") or "",
            stars          = data.get("stargazers_count", 0),
            forks          = data.get("forks_count", 0),
            size_kb        = data.get("size", 0),
            default_branch = data.get("default_branch") or "main",
            topics         = data.get("topics", []),
            last_push      = (data.get("pushed_at") or "")[:10],
            created_at     = (data.get("created_at") or "")[:10],
            archived       = data.get("archived", False),
            html_url       = data.get("html_url", ""),
            open_issues    = data.get("open_issues_count", 0),
        )


@dataclass
class GHBranch:
    org:           str
    repo:          str
    branch:        str
    last_commit:   str
    days_inactive: int
    author:        str
    commit_sha:    str
    repo_url:      str


@dataclass
class SecurityFinding:
    org:      str
    repo:     str
    check:    str
    severity: str    # Critical | High | Medium | Low | Info
    detail:   str
    repo_url: str


@dataclass
class OrgResult:
    org:            str
    members:        List[GHUser]           = field(default_factory=list)
    member_repos:   List[GHRepo]           = field(default_factory=list)
    org_repos:      List[GHRepo]           = field(default_factory=list)
    stale_branches: List[GHBranch]         = field(default_factory=list)
    findings:       List[SecurityFinding]  = field(default_factory=list)
    errors:         List[str]              = field(default_factory=list)
    elapsed:        float                  = 0.0


# ══════════════════════════════════════════════════════════════════════════════
# Token resolution
# ══════════════════════════════════════════════════════════════════════════════

def resolve_token() -> str:
    """env vars first (CI-friendly), gh CLI fallback."""
    for var in ("GITHUB_TOKEN", "GH_TOKEN"):
        tok = os.environ.get(var, "").strip()
        if tok:
            console.print(f"  [green]✓[/]  Token from env [bold cyan]{var}[/]")
            return tok

    if not shutil.which("gh"):
        console.print(Panel(
            "[bold red]✗  No GitHub token found[/]\n\n"
            "  Set [bold cyan]GITHUB_TOKEN[/] or [bold cyan]GH_TOKEN[/] env var,\n"
            "  or install gh CLI: [dim]https://cli.github.com[/] → gh auth login",
            border_style="red", box=box.ROUNDED, expand=False,
        ))
        sys.exit(1)

    try:
        r = subprocess.run(["gh", "auth", "token"],
                           capture_output=True, text=True, timeout=10)
    except subprocess.TimeoutExpired:
        console.print("  [bold red]✗[/]  gh auth token timed out")
        sys.exit(1)

    tok = r.stdout.strip()
    if not tok or r.returncode != 0:
        console.print(Panel(
            "[bold red]✗  gh CLI not authenticated[/]\n\n"
            "  Run [bold cyan]gh auth login[/] and retry.",
            border_style="red", box=box.ROUNDED, expand=False,
        ))
        sys.exit(1)

    console.print("  [green]✓[/]  Token via [bold cyan]gh auth token[/]")
    return tok


# ══════════════════════════════════════════════════════════════════════════════
# GitHub API client
# ══════════════════════════════════════════════════════════════════════════════

class GitHubClient:
    """REST v3 wrapper — auth, pagination, rate-limit back-off, retry."""

    def __init__(self, token: str, timeout: int = 30) -> None:
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization":        f"Bearer {token}",
            "Accept":               "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        })
        self.timeout         = timeout
        self._rate_remaining = 5000
        self._profile_cache: Dict[str, dict] = {}
        self._me:            dict = {}
        self._scopes:        str  = ""
        self._verify()

    def _verify(self) -> None:
        resp = self._get("/user")
        if resp.status_code in (401, 403):
            console.print(
                f"  [bold red]✗[/]  Auth failed ({resp.status_code}). "
                "Check your token or run [bold]gh auth refresh[/]."
            )
            sys.exit(1)
        self._me     = resp.json()
        self._scopes = resp.headers.get("X-OAuth-Scopes", "")
        console.print(
            f"  [green]✓[/]  Authenticated as [bold cyan]{self._me.get('login')}[/]"
            f"  [dim]({self._me.get('name', '')})[/]"
        )
        console.print(f"  [dim]Scopes: {self._scopes or 'fine-grained token'}[/]")
        missing = []
        if self._scopes:
            if "read:org" not in self._scopes and "admin:org" not in self._scopes:
                missing.append("read:org")
            if "admin:org" not in self._scopes:
                missing.append("admin:org (needed for 2FA check)")
        if missing:
            console.print(
                f"  [bold yellow]⚠[/]  Missing scope(s): {', '.join(missing)}\n"
                "  Run [bold]gh auth refresh -s read:org,admin:org[/] for full output."
            )
        console.print()

    def _get(self, path: str, params: Optional[Dict] = None,
             accept: Optional[str] = None) -> requests.Response:
        url     = f"{GITHUB_API}{path}" if path.startswith("/") else path
        backoff = 2
        hdrs    = {"Accept": accept} if accept else {}
        while True:
            try:
                resp = self.session.get(url, params=params,
                                        timeout=self.timeout, headers=hdrs)
            except (requests.ConnectionError, requests.Timeout) as exc:
                console.print(
                    f"  [yellow]⚠[/]  Network error — retrying in {backoff}s "
                    f"({exc.__class__.__name__})"
                )
                time.sleep(backoff)
                backoff = min(backoff * 2, 60)
                continue

            try:
                self._rate_remaining = int(
                    resp.headers.get("X-RateLimit-Remaining", self._rate_remaining)
                )
            except (ValueError, TypeError):
                pass

            if resp.status_code == 429 or (
                resp.status_code == 403 and "rate limit" in resp.text.lower()
            ):
                reset = int(resp.headers.get("X-RateLimit-Reset", time.time() + 60))
                wait  = max(reset - int(time.time()), 1)
                console.print(f"  [bold yellow]⚠[/]  Rate limit — sleeping {wait}s...")
                time.sleep(wait + 2)
                continue

            if resp.status_code >= 500:
                console.print(
                    f"  [yellow]⚠[/]  Server error {resp.status_code} — "
                    f"retrying in {backoff}s"
                )
                time.sleep(backoff)
                backoff = min(backoff * 2, 60)
                continue

            return resp

    def paginate(self, path: str, params: Optional[Dict] = None,
                 accept: Optional[str] = None) -> Iterator[dict]:
        p = {**(params or {}), "per_page": 100, "page": 1}
        while True:
            resp = self._get(path, params=p, accept=accept)
            if resp.status_code == 404:
                return
            if not resp.ok:
                LOG.warning("HTTP %s on %s", resp.status_code, path)
                return
            items = resp.json()
            if not isinstance(items, list) or not items:
                return
            yield from items
            if 'rel="next"' not in resp.headers.get("Link", ""):
                return
            p["page"] += 1

    def get_json(self, path: str, accept: Optional[str] = None) -> Optional[dict]:
        resp = self._get(path, accept=accept)
        return resp.json() if resp.ok else None

    def status_code(self, path: str, accept: Optional[str] = None) -> int:
        return self._get(path, accept=accept).status_code

    def get_user_profile(self, login: str) -> dict:
        """Cached GET /users/{login} — no duplicate calls across orgs."""
        if login not in self._profile_cache:
            self._profile_cache[login] = self.get_json(f"/users/{login}") or {}
        return self._profile_cache[login]

    def check_rate_limit(self) -> None:
        data = self.get_json("/rate_limit") or {}
        rate = data.get("rate", {})
        rem  = rate.get("remaining", self._rate_remaining)
        lim  = rate.get("limit", 5000)
        rst  = rate.get("reset", 0)
        rst_s = datetime.fromtimestamp(rst).strftime("%H:%M:%S") if rst else "?"
        colour = "green" if rem > 1000 else ("yellow" if rem > 200 else "red")
        console.print(
            f"  [green]✓[/]  Rate limit: [bold {colour}]{rem:,}[/] / {lim:,} remaining "
            f"[dim](resets at {rst_s})[/]"
        )
        if rem < 100:
            console.print("  [bold red]⚠  < 100 calls left — scan may be incomplete.[/]")
        console.print()

    @property
    def has_admin_org(self) -> bool:
        return "admin:org" in self._scopes

    @property
    def my_login(self) -> str:
        return self._me.get("login", "")


# ══════════════════════════════════════════════════════════════════════════════
# Core data-fetching
# ══════════════════════════════════════════════════════════════════════════════

def get_default_orgs(client: GitHubClient) -> List[str]:
    """All org logins the authenticated user belongs to."""
    return [item["login"] for item in client.paginate("/user/orgs")]


def get_org_members(client: GitHubClient, org: str) -> List[GHUser]:
    """Single paginate for all members; separate call just to tag admin roles."""
    admin_logins = {
        item.get("login", "")
        for item in client.paginate(f"/orgs/{org}/members", {"role": "admin"})
    }
    members: List[GHUser] = []
    seen: set = set()
    for item in client.paginate(f"/orgs/{org}/members"):
        login = item.get("login", "")
        if login in seen:
            continue
        seen.add(login)
        profile = client.get_user_profile(login) or item
        role    = "admin" if login in admin_logins else "member"
        members.append(GHUser.from_api(profile, role=role))
    return members


def get_member_repos(
    client:        GitHubClient,
    login:         str,
    include_forks: bool,
    since:         Optional[str],
) -> List[GHRepo]:
    """Public repos personally owned by a member login."""
    repos = []
    for item in client.paginate(
        f"/users/{login}/repos", {"type": "public", "sort": "updated"}
    ):
        r = GHRepo.from_api(item)
        if not include_forks and r.fork:
            continue
        if since and r.last_push and r.last_push < since:
            continue
        repos.append(r)
    return repos


def get_org_repos(client: GitHubClient, org: str) -> List[GHRepo]:
    """All repos (public + private) owned directly by the organisation."""
    return [
        GHRepo.from_api(item)
        for item in client.paginate(f"/orgs/{org}/repos", {"type": "all", "sort": "updated"})
    ]


def get_stale_branches(
    client:     GitHubClient,
    org:        str,
    repos:      List[GHRepo],
    stale_days: int,
    workers:    int,
    progress:   Progress,
    task_id,
) -> List[GHBranch]:
    """
    For each org repo, list all branches and check last-commit date.
    Returns branches with no commit activity for > stale_days days.
    Branch commit dates are fetched concurrently per repo.
    """
    cutoff = datetime.now(timezone.utc)
    result: List[GHBranch] = []

    def _check_repo(repo: GHRepo) -> List[GHBranch]:
        found = []
        try:
            for branch in client.paginate(f"/repos/{repo.full_name}/branches"):
                sha = branch.get("commit", {}).get("sha", "")
                if not sha:
                    continue
                commit = client.get_json(f"/repos/{repo.full_name}/commits/{sha}")
                if not commit:
                    continue
                date_str = (
                    commit.get("commit", {}).get("committer", {}).get("date", "")
                )
                if not date_str:
                    continue
                dt   = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
                days = (cutoff - dt).days
                if days > stale_days:
                    author = commit.get("commit", {}).get("author", {}).get("name", "")
                    found.append(GHBranch(
                        org           = org,
                        repo          = repo.name,
                        branch        = branch.get("name", ""),
                        last_commit   = date_str[:10],
                        days_inactive = days,
                        author        = author,
                        commit_sha    = sha[:7],
                        repo_url      = repo.html_url,
                    ))
        except Exception as exc:
            LOG.warning("Branch check failed %s: %s", repo.full_name, exc)
        return found

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_check_repo, r): r for r in repos}
        for fut in as_completed(futures):
            progress.advance(task_id)
            result.extend(fut.result())

    return sorted(result, key=lambda b: (-b.days_inactive, b.repo, b.branch))


# ══════════════════════════════════════════════════════════════════════════════
# Security audit
# ══════════════════════════════════════════════════════════════════════════════

def run_security_audit(
    client:         GitHubClient,
    org:            str,
    members:        List[GHUser],
    org_repos:      List[GHRepo],
    stale_branches: List[GHBranch],
    workers:        int,
) -> List[SecurityFinding]:
    findings: List[SecurityFinding] = []

    # 1. Members without 2FA (requires admin:org scope)
    if client.has_admin_org:
        no_2fa = list(client.paginate(
            f"/orgs/{org}/members", {"filter": "2fa_disabled"}
        ))
        for m in no_2fa:
            findings.append(SecurityFinding(
                org      = org,
                repo     = "— org-level",
                check    = "Member without 2FA",
                severity = "Critical",
                detail   = (
                    f"@{m.get('login')} has 2FA disabled — "
                    "if credentials are compromised the account can be taken over"
                ),
                repo_url = m.get("html_url", ""),
            ))
    else:
        findings.append(SecurityFinding(
            org      = org,
            repo     = "— org-level",
            check    = "2FA check skipped",
            severity = "Info",
            detail   = "Token missing admin:org scope — run: gh auth refresh -s admin:org",
            repo_url = f"https://github.com/orgs/{org}/settings/security",
        ))

    # 2. Per-repo checks (concurrent)
    _SENSITIVE = [
        "infra", "infrastructure", "internal", "private", "secret",
        "credentials", "config", "deploy", "ci", "ops", "prod", "staging",
    ]

    def _check_repo(repo: GHRepo) -> List[SecurityFinding]:
        f: List[SecurityFinding] = []
        url = repo.html_url

        # 2a — No description
        if not (repo.description or "").strip():
            f.append(SecurityFinding(
                org=org, repo=repo.name, check="No Description",
                severity="Low",
                detail="Repo has no description — ownership and purpose are unclear",
                repo_url=url,
            ))

        # 2b — Default branch has no protection rules
        sc = client.status_code(
            f"/repos/{repo.full_name}/branches/{repo.default_branch}/protection"
        )
        if sc == 404:
            f.append(SecurityFinding(
                org=org, repo=repo.name,
                check="No Branch Protection",
                severity="High",
                detail=(
                    f"Default branch '{repo.default_branch}' has no protection rules — "
                    "force-push and direct commits to main are allowed"
                ),
                repo_url=url,
            ))

        # 2c — No license file
        if client.status_code(f"/repos/{repo.full_name}/license") == 404:
            f.append(SecurityFinding(
                org=org, repo=repo.name, check="No License",
                severity="Medium",
                detail="No LICENSE file — legal use of the code is undefined",
                repo_url=url,
            ))

        # 2d — Dependabot vulnerability alerting disabled
        dep_sc = client.status_code(
            f"/repos/{repo.full_name}/vulnerability-alerts",
            accept="application/vnd.github.dorian-preview+json",
        )
        if dep_sc == 404:
            f.append(SecurityFinding(
                org=org, repo=repo.name,
                check="Dependabot Alerts Disabled",
                severity="High",
                detail=(
                    "Dependency vulnerability alerting is OFF — "
                    "known CVEs in dependencies go undetected"
                ),
                repo_url=url,
            ))

        # 2e — Public repo with no topics (classification gap)
        if not repo.private and not repo.topics:
            f.append(SecurityFinding(
                org=org, repo=repo.name, check="No Topics",
                severity="Low",
                detail="Public repo with no topics — difficult to classify or discover",
                repo_url=url,
            ))

        # 2f — Archived repo with open issues (stale noise)
        if repo.archived and repo.open_issues > 0:
            f.append(SecurityFinding(
                org=org, repo=repo.name,
                check="Archived Repo with Open Issues",
                severity="Low",
                detail=(
                    f"Archived but has {repo.open_issues} open issue(s) — "
                    "contributors may not realise the repo is frozen"
                ),
                repo_url=url,
            ))

        # 2g — Public repo with sensitive-sounding name
        name_lower = repo.name.lower()
        hit = next((kw for kw in _SENSITIVE if kw in name_lower), None)
        if not repo.private and hit:
            f.append(SecurityFinding(
                org=org, repo=repo.name,
                check="Potentially Sensitive Public Repo",
                severity="Medium",
                detail=(
                    f"Repo name '{repo.name}' contains '{hit}' but is publicly visible — "
                    "verify this is intentional and contains no secrets"
                ),
                repo_url=url,
            ))

        return f

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_check_repo, r): r for r in org_repos}
        for fut in as_completed(futures):
            findings.extend(fut.result())

    # 3. Org-level: admin-to-member ratio
    admin_ct  = sum(1 for m in members if m.role == "admin")
    member_ct = len(members) - admin_ct
    if admin_ct > 0 and member_ct > 0 and (admin_ct / max(member_ct, 1)) > 0.5:
        findings.append(SecurityFinding(
            org=org, repo="— org-level",
            check="High Admin-to-Member Ratio",
            severity="Medium",
            detail=(
                f"{admin_ct} admins vs {member_ct} members — "
                "too many admins increases the blast radius of a compromised account"
            ),
            repo_url=f"https://github.com/orgs/{org}/people",
        ))

    # 4. Stale branch count summary
    if stale_branches:
        findings.append(SecurityFinding(
            org=org, repo="— org-level",
            check="Stale Branches Present",
            severity="Low",
            detail=(
                f"{len(stale_branches)} stale branch(es) found across org repos — "
                "old branches may contain outdated or insecure code"
            ),
            repo_url=f"https://github.com/{org}",
        ))

    # Sort by severity
    _sev = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Info": 4}
    findings.sort(key=lambda f: (_sev.get(f.severity, 5), f.repo))
    return findings


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════

def _dedupe(items: list, key_fn) -> list:
    seen, out = set(), []
    for x in items:
        k = key_fn(x)
        if k not in seen:
            seen.add(k)
            out.append(x)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# Org processing
# ══════════════════════════════════════════════════════════════════════════════

def process_org(
    client:        GitHubClient,
    org:           str,
    include_forks: bool,
    since:         Optional[str],
    members_only:  bool,
    stale_days:    int,
    run_security:  bool,
    workers:       int,
    progress:      Progress,
) -> OrgResult:
    result  = OrgResult(org=org)
    t_start = time.monotonic()

    console.print(Rule(f"[bold cyan]  {org}  [/]", style="bright_blue"))
    console.print()

    # Step 1 — Members
    with console.status(
        f"  [dim]Fetching members of [bold]{org}[/]...[/]", spinner="dots"
    ):
        try:
            result.members = get_org_members(client, org)
        except Exception as exc:
            result.errors.append(f"members: {exc}")

    admins = sum(1 for m in result.members if m.role == "admin")
    console.print(
        f"  [green]✓[/]  Members — [bold]{len(result.members)}[/] found  "
        f"[dim]({admins} admins · {len(result.members)-admins} members)[/]"
    )
    console.print()

    if members_only:
        result.elapsed = round(time.monotonic() - t_start, 1)
        return result

    # Step 2 — Member personal repos (concurrent)
    console.print("  [dim cyan]Fetching member personal repos (parallel)...[/]")
    task_m = progress.add_task(
        f"[cyan]{org}[/] — member repos", total=len(result.members)
    )

    def _fetch_member(m: GHUser) -> Tuple[List[GHRepo], Optional[str]]:
        progress.update(task_m, description=f"[cyan]{org}[/] — [white]{m.login[:20]}[/]")
        try:
            return get_member_repos(client, m.login, include_forks, since), None
        except Exception as exc:
            return [], f"{m.login}: {exc}"

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_fetch_member, m): m for m in result.members}
        for fut in as_completed(futures):
            repos, err = fut.result()
            result.member_repos.extend(repos)
            if err:
                result.errors.append(err)
            progress.advance(task_m)

    console.print(
        f"  [green]✓[/]  Member repos — [bold]{len(result.member_repos)}[/] across "
        f"[bold]{len(result.members)}[/] members"
    )
    console.print()

    # Step 3 — Org-level repos
    with console.status(
        f"  [dim]Fetching org repos of [bold]{org}[/]...[/]", spinner="dots"
    ):
        try:
            result.org_repos = get_org_repos(client, org)
        except Exception as exc:
            result.errors.append(f"org_repos: {exc}")

    pub_ct  = sum(1 for r in result.org_repos if not r.private)
    priv_ct = sum(1 for r in result.org_repos if r.private)
    console.print(
        f"  [green]✓[/]  Org repos — [bold]{len(result.org_repos)}[/] total  "
        f"[dim]({pub_ct} public · {priv_ct} private)[/]"
    )
    console.print()

    # Step 4 — Stale branches on org repos
    if result.org_repos:
        console.print(
            f"  [dim cyan]Checking stale branches (>{stale_days}d) across "
            f"{len(result.org_repos)} org repo(s)...[/]"
        )
        task_b = progress.add_task(
            f"[cyan]{org}[/] — stale branches", total=len(result.org_repos)
        )
        try:
            result.stale_branches = get_stale_branches(
                client, org, result.org_repos, stale_days, workers, progress, task_b
            )
        except Exception as exc:
            result.errors.append(f"stale_branches: {exc}")

        colour = "red" if result.stale_branches else "green"
        console.print(
            f"  [green]✓[/]  Stale branches — "
            f"[bold {colour}]{len(result.stale_branches)}[/] found "
            f"[dim](>{stale_days} days inactive)[/]"
        )
        console.print()

    # Step 5 — Security audit
    if run_security and result.org_repos:
        console.print("  [dim cyan]Running security audit...[/]")
        try:
            result.findings = run_security_audit(
                client, org, result.members,
                result.org_repos, result.stale_branches, workers
            )
        except Exception as exc:
            result.errors.append(f"security_audit: {exc}")

        crit = sum(1 for f in result.findings if f.severity == "Critical")
        high = sum(1 for f in result.findings if f.severity == "High")
        console.print(
            f"  [green]✓[/]  Security audit — [bold]{len(result.findings)}[/] finding(s)  "
            f"[dim]([bold red]{crit} Critical[/]  [bold yellow]{high} High[/])[/]"
        )
        console.print()

    if result.errors:
        console.print(
            f"  [bold yellow]⚠[/]  {len(result.errors)} error(s)  "
            f"[dim]{result.errors[0][:80]}[/]"
        )
        console.print()

    result.elapsed = round(time.monotonic() - t_start, 1)
    return result


# ══════════════════════════════════════════════════════════════════════════════
# Export
# ══════════════════════════════════════════════════════════════════════════════

def _style_sheet(ws, col_widths: Dict[str, int]) -> None:
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    for cell in ws[1]:
        cell.fill = _HDR; cell.font = _HFNT; cell.border = _BDR
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = _DFNT; cell.border = _BDR
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _colour_by_col(ws, col_idx: int, mapping: Dict) -> None:
    for row in ws.iter_rows(min_row=2):
        fill = mapping.get(str(row[col_idx - 1].value or ""), FILL_WHITE)
        for cell in row:
            cell.fill = fill


def export_all(results: List[OrgResult], outdir: str, fmt: str) -> str:
    os.makedirs(outdir, exist_ok=True)
    base = os.path.join(outdir, f"github_recon_{DATE_FILE}")

    # ── Collect + deduplicate ─────────────────────────────────────────────
    all_members = _dedupe(
        [m for r in results for m in r.members], key_fn=lambda m: m.login
    )
    org_names   = {r.org.lower() for r in results}
    mem_logins  = {m.login.lower() for m in all_members}
    all_m_repos = _dedupe(
        [rp for r in results for rp in r.member_repos], key_fn=lambda rp: rp.html_url
    )
    personal_repos = [
        rp for rp in all_m_repos
        if rp.owner.lower() not in org_names and rp.owner.lower() in mem_logins
    ]
    all_org_repos    = [rp for r in results for rp in r.org_repos]
    all_stale        = [b  for r in results for b  in r.stale_branches]
    all_findings     = [f  for r in results for f  in r.findings]

    # ── Build rows ────────────────────────────────────────────────────────
    summary_rows = [
        {
            "Org":              r.org,
            "Members":          len(r.members),
            "Admins":           sum(1 for m in r.members if m.role == "admin"),
            "Member Repos":     len(r.member_repos),
            "Org Repos":        len(r.org_repos),
            "Org Public Repos": sum(1 for x in r.org_repos if not x.private),
            "Stale Branches":   len(r.stale_branches),
            "Findings":         len(r.findings),
            "Critical Findings": sum(1 for f in r.findings if f.severity == "Critical"),
            "High Findings":    sum(1 for f in r.findings if f.severity == "High"),
            "Elapsed (s)":      r.elapsed,
            "Errors":           len(r.errors),
        }
        for r in results
    ]

    member_rows = [
        {
            "Sno":          i,
            "Login":        m.login,
            "Name":         m.name,
            "Role":         m.role,
            "Email":        m.email,
            "Company":      m.company,
            "Location":     m.location,
            "Public Repos": m.public_repos,
            "Followers":    m.followers,
            "Joined":       m.created_at,
            "Profile URL":  m.profile_url,
        }
        for i, m in enumerate(sorted(all_members, key=lambda m: m.login.lower()), 1)
    ]

    member_repo_rows = [
        {
            "Sno":        i,
            "Owner":      rp.owner,
            "Repo":       rp.name,
            "URL":        rp.html_url,
            "Language":   rp.language,
            "Stars":      rp.stars,
            "Forks":      rp.forks,
            "Last Push":  rp.last_push,
            "Created":    rp.created_at,
            "Fork":       rp.fork,
            "Archived":   rp.archived,
            "Description": rp.description,
        }
        for i, rp in enumerate(
            sorted(personal_repos, key=lambda rp: (rp.owner.lower(), rp.name.lower())), 1
        )
    ]

    org_repo_rows = [
        {
            "Sno":            i,
            "Org":            rp.owner,
            "Repo":           rp.name,
            "URL":            rp.html_url,
            "Visibility":     "Private" if rp.private else "Public",
            "Language":       rp.language,
            "Stars":          rp.stars,
            "Forks":          rp.forks,
            "Open Issues":    rp.open_issues,
            "Default Branch": rp.default_branch,
            "Last Push":      rp.last_push,
            "Created":        rp.created_at,
            "Archived":       rp.archived,
            "Fork":           rp.fork,
            "Topics":         ", ".join(rp.topics),
            "Description":    rp.description,
        }
        for i, rp in enumerate(
            sorted(all_org_repos, key=lambda rp: (rp.owner.lower(), rp.name.lower())), 1
        )
    ]

    stale_rows = [
        {
            "Sno":           i,
            "Org":           b.org,
            "Repo":          b.repo,
            "Branch":        b.branch,
            "Last Commit":   b.last_commit,
            "Days Inactive": b.days_inactive,
            "Last Author":   b.author,
            "Commit SHA":    b.commit_sha,
            "Repo URL":      b.repo_url,
        }
        for i, b in enumerate(all_stale, 1)
    ]

    finding_rows = [
        {
            "Sno":      i,
            "Org":      f.org,
            "Repo":     f.repo,
            "Check":    f.check,
            "Severity": f.severity,
            "Detail":   f.detail,
            "URL":      f.repo_url,
        }
        for i, f in enumerate(all_findings, 1)
    ]

    # ── JSON ──────────────────────────────────────────────────────────────
    if fmt == "json":
        path = f"{base}.json"
        payload = {
            "generated":         datetime.now(timezone.utc).isoformat(),
            "summary":           summary_rows,
            "members":           member_rows,
            "member_repos":      member_repo_rows,
            "org_repos":         org_repo_rows,
            "stale_branches":    stale_rows,
            "security_findings": finding_rows,
        }
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, default=str)
        return os.path.abspath(path)

    # ── CSV ───────────────────────────────────────────────────────────────
    if fmt == "csv":
        for name, rows in [
            ("summary",       summary_rows),
            ("members",       member_rows),
            ("member_repos",  member_repo_rows),
            ("org_repos",     org_repo_rows),
            ("stale_branches", stale_rows),
            ("security",      finding_rows),
        ]:
            pd.DataFrame(rows).to_csv(f"{base}_{name}.csv", index=False)
        return os.path.abspath(outdir)

    # ── XLSX ──────────────────────────────────────────────────────────────
    path   = f"{base}.xlsx"
    sheets = {
        "Summary":        pd.DataFrame(summary_rows),
        "Members":        pd.DataFrame(member_rows),
        "Member Repos":   pd.DataFrame(member_repo_rows),
        "Org Repos":      pd.DataFrame(org_repo_rows),
        "Stale Branches": pd.DataFrame(stale_rows),
        "Security Audit": pd.DataFrame(finding_rows),
    }

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        for sheet, df in sheets.items():
            df.to_excel(w, index=False, sheet_name=sheet)

    wb = load_workbook(path)

    _style_sheet(wb["Summary"],
                 {"A":22,"B":9,"C":8,"D":13,"E":10,"F":16,"G":15,"H":10,"I":17,"J":14,"K":11,"L":7})
    _style_sheet(wb["Members"],
                 {"A":5,"B":20,"C":22,"D":8,"E":26,"F":22,"G":18,"H":12,"I":10,"J":12,"K":38})
    _style_sheet(wb["Member Repos"],
                 {"A":5,"B":18,"C":26,"D":44,"E":12,"F":7,"G":7,"H":12,"I":12,"J":7,"K":9,"L":40})
    _style_sheet(wb["Org Repos"],
                 {"A":5,"B":18,"C":26,"D":44,"E":9,"F":12,"G":7,"H":7,"I":12,"J":16,"K":12,"L":12,"M":9,"N":7,"O":22,"P":40})
    _style_sheet(wb["Stale Branches"],
                 {"A":5,"B":18,"C":26,"D":30,"E":14,"F":14,"G":22,"H":10,"I":44})
    _style_sheet(wb["Security Audit"],
                 {"A":5,"B":18,"C":30,"D":32,"E":10,"F":70,"G":44})

    # Org Repos — grey for archived
    ws_or = wb["Org Repos"]
    arch_idx = list(sheets["Org Repos"].columns).index("Archived") + 1
    for row in ws_or.iter_rows(min_row=2):
        if str(row[arch_idx - 1].value).lower() == "true":
            for cell in row:
                cell.fill = FILL_GREY

    # Stale Branches — red if >365d, amber if >180d
    ws_sb = wb["Stale Branches"]
    for row in ws_sb.iter_rows(min_row=2):
        days = row[5].value or 0    # column F = Days Inactive (0-indexed: 5)
        fill = FILL_RED if days > 365 else (FILL_AMBER if days > 180 else FILL_WHITE)
        for cell in row:
            cell.fill = fill

    # Security Audit — colour by severity
    ws_sec   = wb["Security Audit"]
    sev_idx  = list(sheets["Security Audit"].columns).index("Severity") + 1
    sev_map  = {
        "Critical": FILL_RED,
        "High":     FILL_AMBER,
        "Medium":   FILL_BLUE,
        "Low":      FILL_GREEN,
        "Info":     FILL_GREY,
    }
    _colour_by_col(ws_sec, sev_idx, sev_map)

    wb.save(path)
    return os.path.abspath(path)


# ══════════════════════════════════════════════════════════════════════════════
# Rich UI
# ══════════════════════════════════════════════════════════════════════════════

def print_banner(quiet: bool) -> None:
    if quiet:
        return
    art = (
        "  ██████╗ ██╗████████╗    ██████╗ ███████╗ ██████╗ ██████╗ ███╗   ██╗\n"
        "  ██╔════╝██║╚══██╔══╝    ██╔══██╗██╔════╝██╔════╝██╔═══██╗████╗  ██║\n"
        "  ██║  ███╗██║   ██║      ██████╔╝█████╗  ██║     ██║   ██║██╔██╗ ██║\n"
        "  ██║   ██║██║   ██║      ██╔══██╗██╔══╝  ██║     ██║   ██║██║╚██╗██║\n"
        "  ╚██████╔╝██║   ██║      ██║  ██║███████╗╚██████╗╚██████╔╝██║ ╚████║\n"
        "   ╚═════╝ ╚═╝   ╚═╝      ╚═╝  ╚═╝╚══════╝ ╚═════╝ ╚═════╝ ╚═╝  ╚═══╝"
    )
    console.print()
    console.print(Panel(
        f"[bold bright_cyan]{art}[/]\n\n"
        f"[bold bright_green]   GitHub Org Security & Repo Analyser  v{__version__}[/]"
        f"  [dim white]|[/]  [bold yellow]   Author: {AUTHOR}[/]",
        border_style="bright_blue", box=box.DOUBLE_EDGE, expand=False, padding=(1, 2),
    ))
    console.print()


def make_progress() -> Progress:
    return Progress(
        SpinnerColumn(spinner_name="dots", style="bright_blue"),
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(bar_width=32, style="bright_blue", complete_style="bright_green"),
        TaskProgressColumn(), MofNCompleteColumn(), TimeElapsedColumn(),
        console=console, transient=False,
    )


def print_findings_table(findings: List[SecurityFinding], org: str) -> None:
    if not findings:
        return
    sev_colour = {
        "Critical": "bold red", "High": "bold yellow",
        "Medium": "bold blue",  "Low": "dim green", "Info": "dim",
    }
    tbl = Table(
        title=f"[bold red]Security Findings — {org}[/]  [dim]({len(findings)})[/]",
        box=box.ROUNDED, border_style="red", header_style="bold red",
        show_lines=True, expand=True,
    )
    tbl.add_column("Sno",      style="dim white", justify="right", width=4)
    tbl.add_column("Repo",     style="bold cyan", min_width=18, no_wrap=True)
    tbl.add_column("Check",    min_width=24)
    tbl.add_column("Severity", justify="center", width=10)
    tbl.add_column("Detail",   min_width=32)

    for i, f in enumerate(findings, 1):
        tbl.add_row(
            str(i), f.repo, f.check,
            f"[{sev_colour.get(f.severity, 'white')}]{f.severity}[/]",
            f.detail,
        )
    console.print()
    console.print(tbl)
    console.print()


def print_summary(results: List[OrgResult], report_path: str, wall: float) -> None:
    def fmt_t(s: float) -> str:
        m, sec = divmod(int(s), 60)
        return f"{m}m {sec}s" if m else f"{sec:.0f}s"

    tbl = Table(
        title="[bold white]Recon Summary[/]",
        box=box.ROUNDED, border_style="bright_green",
        header_style="bold bright_green", show_lines=True,
    )
    for col, kw in [
        ("Org",             {"style": "bold cyan", "no_wrap": True}),
        ("Members",         {"justify": "right"}),
        ("Member Repos",    {"justify": "right", "style": "bold green"}),
        ("Org Repos",       {"justify": "right", "style": "bold green"}),
        ("Stale Branches",  {"justify": "right", "style": "bold yellow"}),
        ("Findings",        {"justify": "right", "style": "bold red"}),
        ("Elapsed",         {"justify": "right", "style": "dim white"}),
    ]:
        tbl.add_column(col, **kw)

    total = {"m": 0, "mr": 0, "or": 0, "sb": 0, "f": 0}
    for r in results:
        total["m"]  += len(r.members)
        total["mr"] += len(r.member_repos)
        total["or"] += len(r.org_repos)
        total["sb"] += len(r.stale_branches)
        total["f"]  += len(r.findings)
        tbl.add_row(
            r.org,
            str(len(r.members)),
            str(len(r.member_repos)),
            str(len(r.org_repos)),
            f"[bold {'red' if r.stale_branches else 'green'}]{len(r.stale_branches)}[/]",
            f"[bold {'red' if r.findings else 'green'}]{len(r.findings)}[/]",
            fmt_t(r.elapsed),
        )

    console.print()
    console.print(tbl)
    console.print()
    console.print(Panel(
        f"[bold green]✓  Complete[/]   [dim]·[/]   "
        f"[white]Orgs:[/] [bold]{len(results)}[/]   "
        f"[white]Members:[/] [bold cyan]{total['m']}[/]   "
        f"[white]Member Repos:[/] [bold bright_green]{total['mr']}[/]   "
        f"[white]Org Repos:[/] [bold bright_green]{total['or']}[/]   "
        f"[white]Stale Branches:[/] [bold yellow]{total['sb']}[/]   "
        f"[white]Findings:[/] [bold red]{total['f']}[/]   "
        f"[white]Runtime:[/] [bold cyan]{fmt_t(wall)}[/]",
        border_style="bright_green", box=box.DOUBLE_EDGE, expand=False,
    ))
    if report_path:
        console.print(f"\n  [green]✓[/]  Report → [dim]{report_path}[/]")
    console.print()


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="gitrecon",
        description="GitRecon v4.0 — GitHub Org Security & Repo Analyser",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Token resolution order:\n"
            "  1. GITHUB_TOKEN env var\n"
            "  2. GH_TOKEN env var\n"
            "  3. gh auth token (gh CLI)\n\n"
            "No-args mode (python gitrecon.py):\n"
            "  Auto-discovers all orgs for the logged-in account and runs a full\n"
            "  security audit on each — same as --orgs <all_your_orgs> --security\n\n"
            "Examples:\n"
            "  python gitrecon.py\n"
            "  python gitrecon.py --orgs acme\n"
            "  python gitrecon.py --orgs acme devops --security\n"
            "  python gitrecon.py --orgs acme --stale-days 90\n"
            "  python gitrecon.py --orgs acme --include-forks --since 2024-01-01\n"
            "  python gitrecon.py --orgs acme --output-format json --no-export\n"
            "  python gitrecon.py --orgs acme --members-only\n"
        ),
    )
    tgt = p.add_argument_group("Targets")
    tgt.add_argument(
        "--orgs", "-o", nargs="*", metavar="ORG", default=None,
        help="Org login(s). Omit entirely to auto-discover all orgs for the token.",
    )
    scope = p.add_argument_group("Scope")
    scope.add_argument("--include-forks", dest="include_forks", action="store_true",
                       help="Include forked repos in member results")
    scope.add_argument("--members-only", dest="members_only", action="store_true",
                       help="Members list only — skip all repo and branch enumeration")
    scope.add_argument("--since", default=None, metavar="YYYY-MM-DD",
                       help="Only member repos pushed on or after this date")
    scope.add_argument(
        "--stale-days", dest="stale_days", type=int, default=STALE_DEFAULT, metavar="DAYS",
        help=f"Days of no commits before a branch is stale (default: {STALE_DEFAULT})",
    )
    scope.add_argument("--security", action="store_true",
                       help="Run security audit (always on when --orgs is omitted)")
    out = p.add_argument_group("Output")
    out.add_argument("--outdir", default="github_recon_out", metavar="DIR",
                     help="Output directory (default: github_recon_out)")
    out.add_argument("--output-format", dest="output_format",
                     choices=["xlsx", "csv", "json"], default="xlsx",
                     help="Export format: xlsx (default) | csv | json")
    out.add_argument("--no-export", dest="no_export", action="store_true",
                     help="Skip file export — terminal output only")
    beh = p.add_argument_group("Behaviour")
    beh.add_argument("--workers", type=int, default=8, metavar="N",
                     help="Concurrent workers for repo/branch fetch (default: 8)")
    beh.add_argument("--timeout", type=int, default=30,
                     help="HTTP timeout per request in seconds (default: 30)")
    beh.add_argument("--quiet", action="store_true",
                     help="Suppress banner (CI-friendly)")
    beh.add_argument("-v", "--verbose", action="store_true",
                     help="Debug logging")
    p.add_argument("--version", action="version", version=f"GitRecon v{__version__}")
    return p.parse_args()


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    args = parse_args()
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.WARNING,
        format="%(asctime)s | %(levelname)-8s | %(message)s",
    )

    print_banner(args.quiet)

    token  = resolve_token()
    client = GitHubClient(token, timeout=args.timeout)
    client.check_rate_limit()

    # Org discovery
    auto_mode = args.orgs is None or len(args.orgs) == 0
    if auto_mode:
        console.print("  [dim cyan]No --orgs specified — auto-discovering orgs...[/]")
        with console.status("  [dim]Fetching your organisations...[/]", spinner="dots"):
            orgs = get_default_orgs(client)
        if not orgs:
            console.print(Panel(
                "[bold yellow]⚠  No organisations found for this account.[/]\n"
                "  Ensure your token has [bold]read:org[/] scope.\n"
                "  Run: [bold cyan]gh auth refresh -s read:org[/]",
                border_style="yellow", box=box.ROUNDED, expand=False,
            ))
            sys.exit(0)
        console.print(
            f"  [green]✓[/]  Auto-discovered [bold]{len(orgs)}[/] org(s): "
            f"[dim]{', '.join(orgs)}[/]"
        )
        console.print()
        run_security = True
    else:
        orgs         = args.orgs
        run_security = args.security

    # Pre-flight summary
    flags = []
    if args.include_forks:  flags.append("include-forks")
    if args.members_only:   flags.append("members-only")
    if args.since:          flags.append(f"since={args.since}")
    if run_security:        flags.append("security-audit")
    flags.append(f"stale-days={args.stale_days}")
    console.print(
        f"  [dim cyan]orgs:[/] [bold white]{', '.join(orgs)}[/]  "
        f"[dim cyan]flags:[/] [bold white]{' · '.join(flags)}[/]  "
        f"[dim cyan]format:[/] [bold white]{args.output_format}[/]"
    )
    console.print()

    t_wall   = time.monotonic()
    results: List[OrgResult] = []
    progress = make_progress()

    with progress:
        for org in orgs:
            result = process_org(
                client        = client,
                org           = org,
                include_forks = args.include_forks,
                since         = args.since,
                members_only  = args.members_only,
                stale_days    = args.stale_days,
                run_security  = run_security,
                workers       = args.workers,
                progress      = progress,
            )
            results.append(result)

    # Print security findings per org in terminal
    if run_security:
        for r in results:
            if r.findings:
                print_findings_table(r.findings, r.org)

    # Export
    report_path = ""
    if not args.no_export:
        with console.status(
            f"[bright_blue]Writing {args.output_format.upper()} report...[/]",
            spinner="dots",
        ):
            report_path = export_all(results, args.outdir, args.output_format)

    print_summary(results, report_path, time.monotonic() - t_wall)


if __name__ == "__main__":
    main()